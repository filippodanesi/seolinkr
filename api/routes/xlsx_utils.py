# Copyright (c) 2025-2026 Filippo Danesi. All rights reserved.
"""XLSX utility routes — sheet listing, column preview."""

from __future__ import annotations

from fastapi import APIRouter, File, UploadFile

from api.deps import temp_upload

router = APIRouter(tags=["xlsx-utils"])


@router.post("/xlsx-sheets")
async def xlsx_sheets(file: UploadFile = File(...)):
    """Return sheet names and column headers for each sheet in an uploaded XLSX."""
    from openpyxl import load_workbook

    content = await file.read()

    with temp_upload(content, ".xlsx") as path:
        wb = load_workbook(str(path), read_only=True)
        result = []
        for name in wb.sheetnames:
            ws = wb[name]
            # Read first row as headers
            headers = []
            for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
                headers = [str(c).strip() if c else "" for c in row]
                break
            # Count data rows (sample first 1000)
            row_count = 0
            for _ in ws.iter_rows(min_row=2, max_row=1001, values_only=True):
                row_count += 1
            result.append({
                "name": name,
                "headers": headers,
                "row_count": row_count,
                "is_active": name == wb.active.title,
            })
        wb.close()

    return result
