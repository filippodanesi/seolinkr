# Copyright (c) 2025-2026 Filippo Danesi. All rights reserved.
"""XLSX file parser. Each row with content is treated as a separate section."""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from seo_linker.models import ContentSection
from seo_linker.parsers.base import BaseParser


class XlsxParser(BaseParser):
    def supported_extensions(self) -> list[str]:
        return [".xlsx"]

    def parse(self, file_path: Path) -> list[ContentSection]:
        wb = load_workbook(str(file_path), read_only=True)
        ws = wb.active
        sections: list[ContentSection] = []

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return sections

        # First row is header — find the content column
        header = [str(c).strip().lower() if c else "" for c in rows[0]]
        content_col = None
        for i, h in enumerate(header):
            if h in ("content", "text", "body", "testo", "contenuto", "descrizione"):
                content_col = i
                break
        # Fallback: use the first column with substantial text
        if content_col is None:
            for i, h in enumerate(header):
                content_col = i
                break

        if content_col is None:
            return sections

        for row_idx, row in enumerate(rows[1:], start=1):
            if content_col >= len(row):
                continue
            cell_value = row[content_col]
            if cell_value and str(cell_value).strip():
                # Use first column as heading/identifier if different from content col
                heading = ""
                if content_col != 0 and row[0]:
                    heading = str(row[0]).strip()
                sections.append(
                    ContentSection(
                        text=str(cell_value).strip(),
                        index=row_idx,
                        heading=heading,
                    )
                )

        wb.close()
        return sections
