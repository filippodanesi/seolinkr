# Copyright (c) 2025-2026 Filippo Danesi. All rights reserved.
"""Link Map pipeline — generate strategic internal link recommendations using GSC data."""

from __future__ import annotations

from collections import defaultdict
from collections.abc import Callable
from itertools import combinations
from pathlib import Path
from urllib.parse import urlparse

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from seo_linker.config import Config
from seo_linker.models import LinkMapEntry, LinkMapResult


def run_link_map_pipeline(
    urls: list[str],
    gsc_site: str,
    *,
    output_path: Path | None = None,
    url_pattern: str | None = None,
    days: int = 90,
    min_shared_queries: int = 1,
    config: Config | None = None,
    gsc_client: object | None = None,
    log_fn: Callable[[str], None] | None = None,
) -> LinkMapResult:
    """Generate a strategic internal link map from GSC query overlap.

    Unlike cross_linker.py (which focuses on blog↔blog), this pipeline works
    across ALL page types: PLPs, blogs, info pages, homepage. It's designed
    for strategic planning (Jonathan's Col E use case).

    Steps:
    1. Fetch GSC query-page data for the given URLs
    2. Build query overlap matrix between all URL pairs
    3. Score and prioritize cross-link recommendations
    4. Output as XLSX with source → target → shared queries → priority

    Args:
        urls: List of URLs to analyze (e.g., from an XLSX column or manual input).
        gsc_site: GSC property URL.
        output_path: Where to write the XLSX output.
        url_pattern: Optional regex to filter pages (e.g., "/collections/|/magazine/").
        days: Lookback window for GSC data.
        min_shared_queries: Minimum shared queries to include a recommendation.
        config: SEOLinkr config.
        gsc_client: Pre-built GSC client (optional).
        log_fn: Logging callback.
    """
    import click

    log_fn = log_fn or click.echo
    config = config or Config.load()

    # Step 1: Initialize GSC client
    if gsc_client is None:
        from seo_linker.gsc.client import GSCClient
        gsc_client = GSCClient(
            service_account_path=config.gsc_service_account or None,
            oauth_client_secrets_path=config.gsc_oauth_secrets or None,
            cache_ttl_hours=config.gsc_cache_ttl,
        )

    # Step 2: Fetch GSC data
    log_fn(f"Fetching GSC data for {gsc_site} (last {days} days)...")

    # Get page metrics for all pages
    page_metrics = gsc_client.get_page_metrics(gsc_site, days)
    log_fn(f"  {len(page_metrics)} pages with metrics")

    # Get query-page data — use url_pattern if provided, otherwise build from URL list
    effective_pattern = url_pattern
    if not effective_pattern and urls:
        # Build a pattern that matches the input URLs' paths
        paths = set()
        for url in urls:
            parsed = urlparse(url)
            # Extract directory part (e.g., /collections/ from /collections/bras)
            parts = parsed.path.strip("/").split("/")
            if len(parts) >= 1:
                paths.add(f"/{parts[0]}/")
        if paths:
            effective_pattern = "|".join(paths)

    if effective_pattern:
        log_fn(f"  Fetching queries for pages matching: {effective_pattern}")
        page_queries = gsc_client.get_magazine_queries(gsc_site, effective_pattern, days)
    else:
        log_fn("  Fetching queries for all pages...")
        page_queries = gsc_client.get_magazine_queries(gsc_site, ".", days)

    log_fn(f"  {len(page_queries)} pages with query data")

    # Step 3: Filter to input URLs if provided
    if urls:
        url_set = set(urls)
        filtered = {u: q for u, q in page_queries.items() if u in url_set}
        # Also try without trailing slash / with trailing slash
        for url in urls:
            alt = url.rstrip("/") + "/" if not url.endswith("/") else url.rstrip("/")
            if alt in page_queries and alt not in filtered:
                filtered[url] = page_queries[alt]
        page_queries = filtered if filtered else page_queries
        log_fn(f"  Filtered to {len(page_queries)} of {len(urls)} input URLs")

    if len(page_queries) < 2:
        log_fn("  Not enough pages with query data for cross-link analysis.")
        return LinkMapResult(total_urls=len(urls), total_recommendations=0)

    # Step 4: Build query overlap matrix
    log_fn("Computing query overlap between pages...")

    # Build query sets per page (top 50 queries by impressions)
    query_sets: dict[str, set[str]] = {}
    for url, queries in page_queries.items():
        query_sets[url] = {q.query for q in queries[:50]}

    entries: list[LinkMapEntry] = []

    for url_a, url_b in combinations(query_sets.keys(), 2):
        shared = query_sets[url_a] & query_sets[url_b]
        if len(shared) < min_shared_queries:
            continue

        # Create bidirectional recommendations
        for source, target in [(url_a, url_b), (url_b, url_a)]:
            source_m = page_metrics.get(source)
            target_m = page_metrics.get(target)

            source_imp = source_m.impressions if source_m else 0
            target_imp = target_m.impressions if target_m else 0
            source_pos = source_m.position if source_m else 0.0
            target_pos = target_m.position if target_m else 0.0

            # Score: shared queries × target opportunity × source authority
            score = len(shared) * (1 + target_imp / 10000)
            if 4 <= target_pos <= 15:
                score *= 2.0  # Boost targets in the "striking distance"
            elif target_pos > 0:
                score *= (1 / max(target_pos, 1))
            score = round(score, 3)

            # Priority level
            if len(shared) >= 5 and target_pos <= 15:
                priority = "critical"
            elif len(shared) >= 3 or (len(shared) >= 2 and target_imp >= 1000):
                priority = "high"
            elif len(shared) >= 2:
                priority = "medium"
            else:
                priority = "low"

            # Generate reasoning
            reasoning = _generate_reasoning(source, target, shared, target_pos, target_imp)

            entries.append(LinkMapEntry(
                source_url=source,
                target_url=target,
                shared_queries=sorted(shared)[:10],
                shared_query_count=len(shared),
                source_impressions=source_imp,
                target_impressions=target_imp,
                source_position=source_pos,
                target_position=target_pos,
                relevance_score=score,
                priority=priority,
                reasoning=reasoning,
            ))

    entries.sort(key=lambda e: e.relevance_score, reverse=True)
    log_fn(f"  Found {len(entries)} cross-link recommendations")

    # Step 5: Write output
    if output_path is None:
        output_path = Path("link_map.xlsx")

    _write_link_map_xlsx(entries, output_path, urls, page_metrics, log_fn)
    log_fn(f"\nLink map saved to: {output_path}")

    return LinkMapResult(
        total_urls=len(urls) if urls else len(page_queries),
        total_recommendations=len(entries),
        entries=entries,
        output_path=str(output_path),
    )


def _generate_reasoning(
    source: str, target: str, shared: set[str], target_pos: float, target_imp: int
) -> str:
    """Generate a human-readable reasoning for the link recommendation."""
    parts = []
    target_path = urlparse(target).path

    if 4 <= target_pos <= 10:
        parts.append(f"Target at pos {target_pos:.1f} (quick win)")
    elif 10 < target_pos <= 20:
        parts.append(f"Target at pos {target_pos:.1f} (page 1 push)")
    elif target_pos <= 3:
        parts.append(f"Target at pos {target_pos:.1f} (strong)")

    if target_imp >= 5000:
        parts.append(f"high volume ({target_imp:,} imp)")
    elif target_imp >= 1000:
        parts.append(f"moderate volume ({target_imp:,} imp)")

    sample_queries = sorted(shared)[:3]
    parts.append(f"shared: {', '.join(sample_queries)}")

    return "; ".join(parts)


def _write_link_map_xlsx(
    entries: list[LinkMapEntry],
    output_path: Path,
    input_urls: list[str],
    page_metrics: dict,
    log_fn: Callable[[str], None],
) -> None:
    """Write the link map to an XLSX with two tabs: Recommendations + Summary."""
    wb = Workbook()

    # --- Tab 1: Recommendations ---
    ws = wb.active
    ws.title = "Link Recommendations"

    headers = [
        "Priority", "Source URL", "Target URL", "Shared Queries",
        "Shared Query Count", "Target Impressions", "Target Position",
        "Score", "Reasoning",
    ]
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Priority colors
    priority_fills = {
        "critical": PatternFill(start_color="FF4444", end_color="FF4444", fill_type="solid"),
        "high": PatternFill(start_color="FF8C00", end_color="FF8C00", fill_type="solid"),
        "medium": PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid"),
        "low": PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid"),
    }

    for row_idx, entry in enumerate(entries, 2):
        ws.cell(row=row_idx, column=1, value=entry.priority.upper())
        ws.cell(row=row_idx, column=2, value=entry.source_url)
        ws.cell(row=row_idx, column=3, value=entry.target_url)
        ws.cell(row=row_idx, column=4, value=", ".join(entry.shared_queries))
        ws.cell(row=row_idx, column=5, value=entry.shared_query_count)
        ws.cell(row=row_idx, column=6, value=entry.target_impressions)
        ws.cell(row=row_idx, column=7, value=round(entry.target_position, 1))
        ws.cell(row=row_idx, column=8, value=entry.relevance_score)
        ws.cell(row=row_idx, column=9, value=entry.reasoning)

        # Color priority cell
        fill = priority_fills.get(entry.priority)
        if fill:
            ws.cell(row=row_idx, column=1).fill = fill
            if entry.priority in ("critical",):
                ws.cell(row=row_idx, column=1).font = Font(color="FFFFFF", bold=True)

    # Set column widths
    widths = [12, 50, 50, 60, 15, 18, 15, 10, 60]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # --- Tab 2: Summary by URL ---
    ws2 = wb.create_sheet("Per-URL Summary")
    sum_headers = [
        "URL", "Impressions", "Position", "Outbound Links (recommended)",
        "Inbound Links (recommended)", "Top Targets",
    ]
    for col, header in enumerate(sum_headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font

    # Aggregate per URL
    outbound: dict[str, list[LinkMapEntry]] = defaultdict(list)
    inbound: dict[str, list[LinkMapEntry]] = defaultdict(list)
    for entry in entries:
        outbound[entry.source_url].append(entry)
        inbound[entry.target_url].append(entry)

    all_urls = set(outbound.keys()) | set(inbound.keys())
    if input_urls:
        all_urls |= set(input_urls)

    for row_idx, url in enumerate(sorted(all_urls), 2):
        m = page_metrics.get(url)
        ws2.cell(row=row_idx, column=1, value=url)
        ws2.cell(row=row_idx, column=2, value=m.impressions if m else 0)
        ws2.cell(row=row_idx, column=3, value=round(m.position, 1) if m else 0)
        ws2.cell(row=row_idx, column=4, value=len(outbound.get(url, [])))
        ws2.cell(row=row_idx, column=5, value=len(inbound.get(url, [])))

        # Top 3 targets
        top_targets = sorted(
            outbound.get(url, []),
            key=lambda e: e.relevance_score,
            reverse=True,
        )[:3]
        target_str = ", ".join(
            urlparse(t.target_url).path for t in top_targets
        )
        ws2.cell(row=row_idx, column=6, value=target_str)

    sum_widths = [50, 15, 12, 25, 25, 60]
    for i, w in enumerate(sum_widths, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # --- Tab 3: Jonathan Format (Col E ready) ---
    ws3 = wb.create_sheet("Col E Fill")
    cole_headers = ["URL", "Internal Link Targets (Col E)"]
    for col, header in enumerate(cole_headers, 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font

    for row_idx, url in enumerate(sorted(all_urls), 2):
        ws3.cell(row=row_idx, column=1, value=url)
        targets = sorted(
            outbound.get(url, []),
            key=lambda e: e.relevance_score,
            reverse=True,
        )[:5]
        # Format as "path1, path2, path3"
        target_paths = [urlparse(t.target_url).path for t in targets]
        ws3.cell(row=row_idx, column=2, value=", ".join(target_paths))

    ws3.column_dimensions["A"].width = 50
    ws3.column_dimensions["B"].width = 80

    wb.save(str(output_path))
    wb.close()
