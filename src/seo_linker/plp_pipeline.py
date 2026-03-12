# Copyright (c) 2025-2026 Filippo Danesi. All rights reserved.
"""PLP batch pipeline — inject internal links into PLP HTML content blocks."""

from __future__ import annotations

from collections.abc import Callable
from pathlib import Path

from openpyxl import load_workbook

from seo_linker.config import Config
from seo_linker.linking.plp_linker import link_plp_html
from seo_linker.matching.prefilter import prefilter_pages
from seo_linker.models import (
    ContentSection,
    PLPBatchResult,
    PLPLinkingResult,
    PLPRow,
    TargetPage,
)
from seo_linker.parsers.plp_xlsx_parser import parse_plp_xlsx
from seo_linker.sitemap.enricher import enrich_pages
from seo_linker.sitemap.fetcher import fetch_sitemap


def run_plp_pipeline(
    input_path: Path,
    sitemap_urls: list[str],
    *,
    output_path: Path | None = None,
    sheet_name: str | None = None,
    url_col: str | None = None,
    content_col: str | None = None,
    keyword_col: str | None = None,
    related_kw_col: str | None = None,
    max_links: int = 5,
    top_n: int = 25,
    model: str | None = None,
    config: Config | None = None,
    gsc_site: str | None = None,
    brand_guidelines: str | None = None,
    gsc_client: object | None = None,
    log_fn: Callable[[str], None] | None = None,
) -> PLPBatchResult:
    """Execute the PLP batch internal linking pipeline.

    1. Parse XLSX to extract PLP rows (URL + HTML content)
    2. Fetch & enrich sitemaps (shared across all rows)
    3. For each row: prefilter candidates → Claude HTML link injection
    4. Write results back to XLSX
    """
    import click

    log_fn = log_fn or click.echo
    config = config or Config.load()
    model = model or config.default_model

    if not config.api_key:
        raise ValueError(
            "Anthropic API key not set. Run: seo-linker config --api-key YOUR_KEY"
        )

    # Step 1: Parse XLSX
    log_fn(f"Parsing PLP XLSX: {input_path.name}...")
    try:
        plp_rows = parse_plp_xlsx(
            input_path,
            sheet_name=sheet_name,
            url_col=url_col,
            content_col=content_col,
            keyword_col=keyword_col,
            related_kw_col=related_kw_col,
        )
    except Exception as e:
        import traceback
        log_fn(f"  PARSE ERROR: {e}")
        log_fn(traceback.format_exc())
        raise
    log_fn(f"  Found {len(plp_rows)} PLP rows with content")

    if not plp_rows:
        return PLPBatchResult(total_rows=0, succeeded=0, failed=0)

    # Step 2: Fetch & enrich sitemaps
    pages: list[TargetPage] = []
    for sitemap_url in sitemap_urls:
        log_fn(f"Fetching sitemap: {sitemap_url}...")
        sitemap_pages = fetch_sitemap(sitemap_url)
        log_fn(f"  Found {len(sitemap_pages)} URLs")
        pages.extend(sitemap_pages)

    # Deduplicate
    seen: set[str] = set()
    pages = [p for p in pages if p.url not in seen and not seen.add(p.url)]

    # Filter product pages
    before = len(pages)
    pages = [p for p in pages if not p.url.rstrip("/").endswith(".html")]
    if before != len(pages):
        log_fn(f"  Filtered {before - len(pages)} product pages, {len(pages)} remaining")

    # Enrich with metadata
    log_fn("Enriching pages with metadata...")
    pages = enrich_pages(pages, config.cache_ttl_hours)
    log_fn(f"  Enriched {sum(1 for p in pages if p.title)}/{len(pages)} pages")

    # GSC enrichment
    if gsc_site:
        if gsc_client is None:
            from seo_linker.gsc.client import GSCClient
            gsc_client = GSCClient(
                service_account_path=config.gsc_service_account or None,
                oauth_client_secrets_path=config.gsc_oauth_secrets or None,
                cache_ttl_hours=config.gsc_cache_ttl,
            )
        pages = gsc_client.enrich_candidates(pages, gsc_site)
        log_fn(f"  GSC: {sum(1 for p in pages if p.impressions > 0)}/{len(pages)} with metrics")

    # Step 3: Process each PLP row
    log_fn(f"\nProcessing {len(plp_rows)} PLPs...")
    row_results: list[PLPLinkingResult] = []
    succeeded = 0
    failed = 0
    total_links = 0

    for i, plp_row in enumerate(plp_rows):
        log_fn(f"  [{i + 1}/{len(plp_rows)}] {plp_row.url}...")

        try:
            # Create a content section from the HTML for prefiltering
            section = ContentSection(
                text=_html_to_text(plp_row.content_html),
                index=0,
                heading=plp_row.target_keyword or plp_row.url,
            )

            # Prefilter candidates for this specific PLP
            candidates = prefilter_pages(
                [section], pages, top_n, config.embedding_model,
            )

            # Remove self-link
            candidates = [c for c in candidates if c.url != plp_row.url]

            if not candidates:
                log_fn(f"    No candidates found, skipping")
                row_results.append(PLPLinkingResult(
                    row_index=plp_row.row_index,
                    url=plp_row.url,
                    original_html=plp_row.content_html,
                    linked_html=plp_row.content_html,
                ))
                succeeded += 1
                continue

            # Claude link injection
            linked_html, insertions = link_plp_html(
                html_content=plp_row.content_html,
                candidate_pages=candidates,
                api_key=config.api_key,
                model=model,
                max_links=max_links,
                current_url=plp_row.url,
                brand_guidelines=brand_guidelines,
                target_keyword=plp_row.target_keyword or None,
                related_keywords=plp_row.related_keywords or None,
            )

            row_results.append(PLPLinkingResult(
                row_index=plp_row.row_index,
                url=plp_row.url,
                original_html=plp_row.content_html,
                linked_html=linked_html,
                insertions=insertions,
            ))
            total_links += len(insertions)
            succeeded += 1
            log_fn(f"    Inserted {len(insertions)} links")

        except Exception as e:
            import traceback
            log_fn(f"    ERROR: {e}")
            log_fn(f"    {traceback.format_exc()}")
            row_results.append(PLPLinkingResult(
                row_index=plp_row.row_index,
                url=plp_row.url,
                original_html=plp_row.content_html,
                linked_html=plp_row.content_html,
            ))
            failed += 1

    # Step 4: Write results back to XLSX
    if output_path is None:
        output_path = input_path.with_name(f"{input_path.stem}_linked{input_path.suffix}")

    _write_plp_results(input_path, output_path, plp_rows, row_results, content_col)
    log_fn(f"\nOutput: {output_path}")
    log_fn(f"Summary: {succeeded}/{len(plp_rows)} succeeded, {total_links} links inserted")

    return PLPBatchResult(
        total_rows=len(plp_rows),
        succeeded=succeeded,
        failed=failed,
        total_links_inserted=total_links,
        row_results=row_results,
        output_path=str(output_path),
    )


def _html_to_text(html: str) -> str:
    """Strip HTML tags to get plain text for embedding."""
    import re
    text = re.sub(r"<[^>]+>", " ", html)
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def _write_plp_results(
    input_path: Path,
    output_path: Path,
    plp_rows: list[PLPRow],
    row_results: list[PLPLinkingResult],
    content_col: str | None,
) -> None:
    """Write linked HTML back to the XLSX, adding a 'Linked Content' column."""
    wb = load_workbook(str(input_path))
    ws = wb.active

    # Find the content column index
    header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    header_lower = [str(c).strip().lower() if c else "" for c in header_row]

    if content_col:
        from seo_linker.parsers.plp_xlsx_parser import _col_letter_to_index
        content_idx = _col_letter_to_index(content_col.upper())
    else:
        # Find by matching the column name from the first PLPRow
        content_idx = None
        if plp_rows and plp_rows[0].column_name:
            target = plp_rows[0].column_name.lower()
            for i, h in enumerate(header_lower):
                if h == target:
                    content_idx = i
                    break
        if content_idx is None:
            # Fallback: search known headers
            from seo_linker.parsers.plp_xlsx_parser import _CONTENT_HEADERS
            for i, h in enumerate(header_lower):
                if h in _CONTENT_HEADERS:
                    content_idx = i
                    break

    # Add "Linked Content" column
    linked_col = len(header_row) + 1
    ws.cell(row=1, column=linked_col, value="Linked Content")

    # Build result lookup by row index
    result_map = {r.row_index: r for r in row_results}

    for plp_row in plp_rows:
        result = result_map.get(plp_row.row_index)
        if result and result.linked_html != result.original_html:
            ws.cell(row=plp_row.row_index, column=linked_col, value=result.linked_html)
        else:
            # No changes — copy original
            ws.cell(row=plp_row.row_index, column=linked_col, value=plp_row.content_html)

    wb.save(str(output_path))
    wb.close()
