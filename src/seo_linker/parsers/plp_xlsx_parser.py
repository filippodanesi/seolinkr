# Copyright (c) 2025-2026 Filippo Danesi. All rights reserved.
"""PLP XLSX parser — extracts URL + HTML content columns for PLP batch processing."""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from seo_linker.models import PLPRow


# Column header patterns for auto-detection (case-insensitive)
_URL_HEADERS = {"url", "page", "page url", "landing page", "address"}
_CONTENT_HEADERS = {
    "bottom seo text", "seo text", "bottom text", "category text",
    "description", "content", "text", "body", "html", "html content",
    "introduction text", "intro text",
}
_KEYWORD_HEADERS = {"target keyword", "main keyword", "keyword", "focus keyword"}
_RELATED_KW_HEADERS = {"related keywords", "secondary keywords", "lsi keywords"}


def _safe_get(row: tuple, idx: int):
    """Safely get a value from a row tuple (openpyxl read_only rows can vary in length)."""
    if idx < len(row):
        return row[idx]
    return None


def parse_plp_xlsx(
    file_path: Path,
    sheet_name: str | None = None,
    url_col: str | None = None,
    content_col: str | None = None,
    keyword_col: str | None = None,
    related_kw_col: str | None = None,
) -> list[PLPRow]:
    """Parse an XLSX file to extract PLP rows with URL and HTML content.

    Auto-detects columns by header name, or uses explicit column letters.

    Args:
        file_path: Path to the XLSX file.
        sheet_name: Sheet to read (default: active sheet).
        url_col: Column letter for URL (e.g., "B"). Auto-detected if None.
        content_col: Column letter for HTML content (e.g., "U"). Auto-detected if None.
        keyword_col: Column letter for target keyword. Auto-detected if None.
        related_kw_col: Column letter for related keywords. Auto-detected if None.

    Returns:
        List of PLPRow with row_index, url, content_html, and keyword metadata.
    """
    wb = load_workbook(str(file_path), read_only=True)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active

    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return []

    # Build header index — pad short rows with empty strings
    header_raw = rows[0]
    header = [str(c).strip().lower() if c else "" for c in header_raw]

    # Auto-detect or resolve column indices
    url_idx = _resolve_col(header, url_col, _URL_HEADERS)
    content_idx = _resolve_col(header, content_col, _CONTENT_HEADERS)
    kw_idx = _resolve_col(header, keyword_col, _KEYWORD_HEADERS)
    related_idx = _resolve_col(header, related_kw_col, _RELATED_KW_HEADERS)

    if url_idx is None:
        raise ValueError(
            f"Could not find URL column. Headers: {header}. "
            "Use --url-col to specify explicitly."
        )
    if content_idx is None:
        raise ValueError(
            f"Could not find content column. Headers: {header}. "
            "Use --content-col to specify explicitly."
        )

    content_header_val = _safe_get(header_raw, content_idx)
    content_header_name = str(content_header_val).strip() if content_header_val else ""

    plp_rows: list[PLPRow] = []
    for row_idx, row in enumerate(rows[1:], start=2):  # Excel row 2+
        url_val = _safe_get(row, url_idx)
        content_val = _safe_get(row, content_idx)

        if not url_val or not content_val:
            continue

        url_str = str(url_val).strip()
        content_str = str(content_val).strip()

        if not url_str or not content_str:
            continue

        kw_val = _safe_get(row, kw_idx) if kw_idx is not None else None
        kw_str = str(kw_val).strip() if kw_val else ""

        related_val = _safe_get(row, related_idx) if related_idx is not None else None
        related_str = str(related_val).strip() if related_val else ""

        plp_rows.append(PLPRow(
            row_index=row_idx,
            url=url_str,
            content_html=content_str,
            column_name=content_header_name,
            target_keyword=kw_str,
            related_keywords=related_str,
        ))

    return plp_rows


def _resolve_col(
    header: list[str],
    explicit_col: str | None,
    auto_detect_headers: set[str],
) -> int | None:
    """Resolve a column index from explicit letter or auto-detection."""
    if explicit_col:
        # Convert letter(s) to 0-based index (A=0, B=1, ..., Z=25, AA=26)
        return _col_letter_to_index(explicit_col.upper())

    # Auto-detect by matching header names
    for i, h in enumerate(header):
        if h in auto_detect_headers:
            return i

    return None


def _col_letter_to_index(col: str) -> int:
    """Convert Excel column letter to 0-based index."""
    result = 0
    for char in col:
        result = result * 26 + (ord(char) - ord("A") + 1)
    return result - 1
