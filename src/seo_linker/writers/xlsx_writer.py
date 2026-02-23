"""XLSX writer — adds a column with the linked version of each content row."""

from __future__ import annotations

import re
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font

from seo_linker.models import LinkingResult
from seo_linker.writers.base import BaseWriter

LINK_RE = re.compile(r"\[([^\]]+)\]\(([^)]+)\)")


class XlsxWriter(BaseWriter):
    def write(self, result: LinkingResult, input_path: Path, output_path: Path) -> None:
        wb = load_workbook(str(input_path))
        ws = wb.active

        # Find or create "Linked Content" column
        header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        linked_col = len(header_row) + 1
        ws.cell(row=1, column=linked_col, value="Linked Content")

        # The linked_text contains all sections joined. Split back by double newline.
        linked_sections = result.linked_text.split("\n\n") if result.linked_text else []

        # Find content column to match rows
        header_lower = [str(c).strip().lower() if c else "" for c in header_row]
        content_col = None
        for i, h in enumerate(header_lower):
            if h in ("content", "text", "body", "testo", "contenuto", "descrizione"):
                content_col = i
                break
        if content_col is None:
            content_col = 0

        section_idx = 0
        for row_idx in range(2, ws.max_row + 1):
            cell_value = ws.cell(row=row_idx, column=content_col + 1).value
            if cell_value and str(cell_value).strip():
                if section_idx < len(linked_sections):
                    linked_text = linked_sections[section_idx]
                    # Write plain text version (XLSX cells can't have inline markdown links
                    # in a single cell easily; write the markdown version)
                    ws.cell(row=row_idx, column=linked_col, value=linked_text)
                    section_idx += 1

        wb.save(str(output_path))
        wb.close()
